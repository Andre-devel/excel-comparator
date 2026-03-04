"""
CAMADA DE GERAÇÃO DE RELATÓRIO
Gera o arquivo .xlsx de saída com:
  - Aba principal: dados do Arquivo 1 com células divergentes em fundo vermelho.
  - Aba "Resumo": estatísticas e lista detalhada de todas as divergências.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers
from typing import Optional

from comparator import ComparisonResult


# Cores
RED_FILL = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
HEADER_FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
HEADER2_FILL = PatternFill(start_color="2D4A7A", end_color="2D4A7A", fill_type="solid")
SUMMARY_HEADER_FILL = PatternFill(start_color="2D6A4F", end_color="2D6A4F", fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color="F7F9FC", end_color="F7F9FC", fill_type="solid")
STAT_FILL = PatternFill(start_color="EBF4FF", end_color="EBF4FF", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
MISSING_FILL = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")

WHITE_FONT = Font(name="Calibri", color="FFFFFF", bold=True, size=11)
BOLD_FONT = Font(name="Calibri", bold=True, size=11)
NORMAL_FONT = Font(name="Calibri", size=10)
RED_FONT = Font(name="Calibri", bold=True, color="CC0000", size=10)

THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)


def generate_report(
    df1: pd.DataFrame,
    result: ComparisonResult,
    output_path: str,
    file1_name: str = "Arquivo 1",
    file2_name: str = "Arquivo 2",
) -> str:
    """
    Gera o arquivo Excel de saída com divergências destacadas.

    Args:
        df1: DataFrame do Arquivo 1 (base para os dados)
        result: Resultado da comparação (ComparisonResult)
        output_path: Caminho onde o arquivo .xlsx será salvo
        file1_name: Nome descritivo do Arquivo 1 (para o Resumo)
        file2_name: Nome descritivo do Arquivo 2 (para o Resumo)

    Returns:
        Caminho do arquivo gerado.
    """
    wb = Workbook()

    _build_main_sheet(wb, df1, result)
    _build_summary_sheet(wb, result, file1_name, file2_name)

    wb.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# PRIVADO
# ---------------------------------------------------------------------------

def _build_main_sheet(wb: Workbook, df1: pd.DataFrame, result: ComparisonResult):
    """Constrói a aba principal com dados do Arquivo 1 e células em vermelho."""
    ws = wb.active
    ws.title = "Comparação"

    columns = list(df1.columns)
    col_count = len(columns)

    # Mapa rápido: (row_number, col_name) -> True para acesso O(1)
    divergent_cells: set[tuple[int, str]] = {
        (d.row_number, d.column_name)
        for d in result.divergences
        if d.row_number > 0
    }

    # Cabeçalho
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    ws.row_dimensions[1].height = 30

    # Linhas de dados
    for row_idx, (_, row) in enumerate(df1.iterrows(), start=1):
        excel_row = row_idx + 1  # +1 pelo cabeçalho
        is_alt = row_idx % 2 == 0

        for col_idx, col_name in enumerate(columns, start=1):
            value = row[col_name]
            cell = ws.cell(row=excel_row, column=col_idx, value=value)

            if (row_idx, col_name) in divergent_cells:
                cell.fill = RED_FILL
                cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
            else:
                if is_alt:
                    cell.fill = ALT_ROW_FILL
                cell.font = NORMAL_FONT

            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = THIN_BORDER

        ws.row_dimensions[excel_row].height = 18

    # Ajuste de largura das colunas
    for col_idx, col_name in enumerate(columns, start=1):
        letter = get_column_letter(col_idx)
        max_len = max(
            len(str(col_name)),
            *[len(str(df1.iloc[i][col_name])) for i in range(min(100, len(df1)))]
        )
        ws.column_dimensions[letter].width = min(max(max_len + 4, 12), 50)

    # Freeze do cabeçalho
    ws.freeze_panes = "A2"

    # Legenda
    legend_row = len(df1) + 3
    ws.cell(row=legend_row, column=1, value="🔴 Célula em vermelho = divergência encontrada").font = RED_FONT


def _build_summary_sheet(
    wb: Workbook,
    result: ComparisonResult,
    file1_name: str,
    file2_name: str,
):
    """Constrói a aba Resumo com estatísticas e lista detalhada de divergências."""
    ws = wb.create_sheet(title="Resumo")

    # ---- ESTATÍSTICAS ----
    stats = [
        ("Total de linhas comparadas", result.total_rows_compared),
        ("Total de divergências encontradas", result.total_divergences),
        ("Linhas com ao menos uma divergência", len(result.divergent_rows)),
    ]

    ws.cell(row=1, column=1, value="📊 RESUMO DA COMPARAÇÃO").font = Font(
        name="Calibri", bold=True, size=14, color="1E3A5F"
    )
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="left")
    ws.merge_cells("A1:D1")

    ws.cell(row=2, column=1, value=f"Arquivo 1: {file1_name}").font = Font(
        name="Calibri", italic=True, size=10, color="555555"
    )
    ws.cell(row=3, column=1, value=f"Arquivo 2: {file2_name}").font = Font(
        name="Calibri", italic=True, size=10, color="555555"
    )

    for i, (label, value) in enumerate(stats, start=5):
        label_cell = ws.cell(row=i, column=1, value=label)
        label_cell.font = BOLD_FONT
        label_cell.fill = STAT_FILL
        label_cell.border = THIN_BORDER
        label_cell.alignment = Alignment(horizontal="left", vertical="center")

        value_cell = ws.cell(row=i, column=2, value=value)
        value_cell.font = Font(name="Calibri", bold=True, size=12, color="1E3A5F")
        value_cell.fill = STAT_FILL
        value_cell.border = THIN_BORDER
        value_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[i].height = 22

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20

    # ---- TABELA DE DIVERGÊNCIAS ----
    header_row = 10
    headers = ["Linha", "Coluna", f"Valor — {file1_name}", f"Valor — {file2_name}"]
    header_widths = [10, 30, 40, 40]

    ws.cell(row=header_row - 1, column=1, value="📋 DETALHAMENTO DAS DIVERGÊNCIAS").font = Font(
        name="Calibri", bold=True, size=12, color="1E3A5F"
    )

    for col_idx, (header, width) in enumerate(zip(headers, header_widths), start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.fill = SUMMARY_HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[header_row].height = 28

    if not result.divergences:
        no_div_cell = ws.cell(
            row=header_row + 1, column=1,
            value="✅ Nenhuma divergência encontrada. Os arquivos são idênticos nas colunas comparadas."
        )
        no_div_cell.font = Font(name="Calibri", bold=True, color="2D6A4F", size=11)
        no_div_cell.alignment = Alignment(horizontal="left")
        ws.merge_cells(f"A{header_row + 1}:D{header_row + 1}")
    else:
        for i, div in enumerate(result.divergences, start=1):
            excel_row = header_row + i
            is_alt = i % 2 == 0
            row_fill = ALT_ROW_FILL if is_alt else PatternFill()

            data = [
                str(div.row_number) if div.row_number > 0 else "N/A",
                div.column_name,
                div.value_file1,
                div.value_file2,
            ]

            for col_idx, value in enumerate(data, start=1):
                cell = ws.cell(row=excel_row, column=col_idx, value=value)
                cell.font = NORMAL_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(
                    horizontal="center" if col_idx == 1 else "left",
                    vertical="center",
                    wrap_text=True,
                )
                if is_alt:
                    cell.fill = ALT_ROW_FILL

            ws.row_dimensions[excel_row].height = 16

    ws.freeze_panes = f"A{header_row + 1}"


# ---------------------------------------------------------------------------
# RELATÓRIO DE CONFERÊNCIA (junção lado a lado)
# ---------------------------------------------------------------------------

def generate_merge_report(
    df1: pd.DataFrame,
    df2: pd.DataFrame,
    output_path: str,
    file1_name: str = "Arquivo 1",
    file2_name: str = "Arquivo 2",
    primary_key: Optional[str] = None,
    ignore_columns: Optional[list[str]] = None,
) -> str:
    """
    Gera um .xlsx com os dois arquivos lado a lado para conferência manual.

    Estrutura de colunas por campo:
        {col} — Arq. 1  |  {col} — Arq. 2  |  (separador vazio)

    Células com valores divergentes são destacadas em amarelo.
    """
    ignore_set = set(ignore_columns or [])
    columns = [c for c in df1.columns if c not in ignore_set]

    wb = Workbook()
    ws = wb.active
    ws.title = "Conferência"

    # ---- Linha de cabeçalho de arquivo (linha 1: grupos) ----
    # ---- Linha de cabeçalho de coluna (linha 2: nomes detalhados) ----
    col_map: dict[str, tuple[int, int]] = {}  # col -> (excel_col_arq1, excel_col_arq2)
    excel_col = 1

    for col in columns:
        ec1, ec2 = excel_col, excel_col + 1

        # Linha 1: nome do arquivo como grupo
        c = ws.cell(row=1, column=ec1, value=file1_name)
        c.fill = HEADER_FILL
        c.font = WHITE_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER

        c = ws.cell(row=1, column=ec2, value=file2_name)
        c.fill = HEADER2_FILL
        c.font = WHITE_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER

        # Linha 2: nome da coluna
        for ec, fill in [(ec1, HEADER_FILL), (ec2, HEADER2_FILL)]:
            c = ws.cell(row=2, column=ec, value=col)
            c.fill = fill
            c.font = Font(name="Calibri", color="FFFFFF", bold=True, size=10)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = THIN_BORDER

        # Coluna separadora: estreita e vazia
        sep_letter = get_column_letter(excel_col + 2)
        ws.column_dimensions[sep_letter].width = 2

        col_map[col] = (ec1, ec2)
        excel_col += 3

    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 28
    ws.freeze_panes = "A3"

    # ---- Linhas de dados ----
    if primary_key and primary_key in df1.columns and primary_key in df2.columns:
        df1_idx = df1.set_index(primary_key)
        df2_idx = df2.set_index(primary_key)
        # Ordem: chaves do arq1 primeiro, depois exclusivas do arq2
        all_keys = list(df1_idx.index) + [k for k in df2_idx.index if k not in df1_idx.index]

        for row_offset, key in enumerate(all_keys):
            excel_row = row_offset + 3
            has1 = key in df1_idx.index
            has2 = key in df2_idx.index
            row1 = df1_idx.loc[key] if has1 else None
            row2 = df2_idx.loc[key] if has2 else None
            is_alt = row_offset % 2 == 0

            for col in columns:
                ec1, ec2 = col_map[col]
                v1 = str(row1[col]) if has1 and col in df1_idx.columns else ""
                v2 = str(row2[col]) if has2 and col in df2_idx.columns else ""

                lbl1 = v1 if has1 else "<ausente>"
                lbl2 = v2 if has2 else "<ausente>"

                c1 = ws.cell(row=excel_row, column=ec1, value=lbl1)
                c2 = ws.cell(row=excel_row, column=ec2, value=lbl2)

                if not has1 or not has2:
                    c1.fill = MISSING_FILL
                    c2.fill = MISSING_FILL
                elif v1 != v2:
                    c1.fill = YELLOW_FILL
                    c2.fill = YELLOW_FILL
                elif is_alt:
                    c1.fill = ALT_ROW_FILL
                    c2.fill = ALT_ROW_FILL

                for c in (c1, c2):
                    c.font = NORMAL_FONT
                    c.alignment = Alignment(horizontal="left", vertical="center")
                    c.border = THIN_BORDER

            ws.row_dimensions[excel_row].height = 18
    else:
        total_rows = max(len(df1), len(df2))
        for row_offset in range(total_rows):
            excel_row = row_offset + 3
            has1 = row_offset < len(df1)
            has2 = row_offset < len(df2)
            is_alt = row_offset % 2 == 0

            for col in columns:
                ec1, ec2 = col_map[col]
                v1 = str(df1.iloc[row_offset][col]) if has1 else ""
                v2 = str(df2.iloc[row_offset][col]) if has2 else ""

                c1 = ws.cell(row=excel_row, column=ec1, value=v1 if has1 else "<ausente>")
                c2 = ws.cell(row=excel_row, column=ec2, value=v2 if has2 else "<ausente>")

                if not has1 or not has2:
                    c1.fill = MISSING_FILL
                    c2.fill = MISSING_FILL
                elif v1 != v2:
                    c1.fill = YELLOW_FILL
                    c2.fill = YELLOW_FILL
                elif is_alt:
                    c1.fill = ALT_ROW_FILL
                    c2.fill = ALT_ROW_FILL

                for c in (c1, c2):
                    c.font = NORMAL_FONT
                    c.alignment = Alignment(horizontal="left", vertical="center")
                    c.border = THIN_BORDER

            ws.row_dimensions[excel_row].height = 18

    # ---- Largura das colunas de dados ----
    for col in columns:
        ec1, ec2 = col_map[col]
        sample = [str(df1.iloc[i][col]) for i in range(min(100, len(df1)))]
        max_len = max(len(col), max((len(v) for v in sample), default=0))
        width = min(max(max_len + 2, 12), 40)
        ws.column_dimensions[get_column_letter(ec1)].width = width
        ws.column_dimensions[get_column_letter(ec2)].width = width

    wb.save(output_path)
    return output_path
